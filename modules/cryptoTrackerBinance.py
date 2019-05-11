from decimal import Decimal
from shutil import copyfile
from openpyxl import load_workbook
from modules.cryptoTrackerUtil import *


def bin_deposit_coin(fn, cur, con):
    # fix Binance bug that downloads file as csv
    if fn[1].split('.')[1] == 'csv':
        copyfile(fn[2], fn[2].replace('csv', 'xlsx'))
    wb = load_workbook(fn[2].replace('csv', 'xlsx'))
    # initialize arrays
    coin_to_exchange = []
    coin_tx_list = {}
    ws = wb['sheet1']
    # print (ws.min_row)
    # print (ws.max_row)
    for row in ws.iter_rows(min_row=2):
        # print (row[0].value)
        date_time = str(row[0].value)
        coin = str(row[1].value) + "_Bin"
        # print (coin)
        add_column_if_not_in_table(coin, cur, con)
        tx_id = row[5].value.replace('0x', '')
        qty = row[2].value
        cur.execute("INSERT INTO crypto (filename, datetime, type, %s, tx_id) VALUES (?, ?, ?, ?, ?);" % coin,
                    (fn[1], date_time, 'xfer', qty, tx_id))
        con.commit()

        coin_addr = str(row[4].value)
        if "BTC" in coin:
            coin_tx_list[tx_id] = coinObject(Decimal(qty) * SATOSHI, fn[1], coin, coin_addr)

        else:
            coin_tx_list[tx_id] = coinObject(Decimal(qty), fn[1], coin, coin_addr)
        coin_to_exchange.append(tx_id)
    return [coin_tx_list, coin_to_exchange]

def bin_widthdraw_coin(fn, cur, con):
    # fix Binance bug that downloads file as csv
    if fn[1].split('.')[1] == 'csv':
        copyfile(fn[2], fn[2].replace('csv', 'xlsx'))
    wb = load_workbook(fn[2].replace('csv', 'xlsx'))
    # initialize arrays
    coin_from_exchange = []
    coin_tx_list = {}
    ws = wb['sheet1']
    # print (ws.min_row)
    # print (ws.max_row)
    for row in ws.iter_rows(min_row=2):
        # print (row[0].value)
        coin_addr = str(row[4].value).replace('0x', '')
        date_time = str(row[0].value)
        coin =  str(row[1].value) + "_Bin"
        # print (coin)
        add_column_if_not_in_table(coin, cur, con)
        wallet = str(row[1].value).upper() + '_' + coin_addr[0:5]
        add_column_if_not_in_table(wallet, cur, con)
        tx_id = row[5].value.replace('0x', '')
        qty = Decimal(row[2].value) * -1
        fee = Decimal(row[3].value) * -1
        cur.execute("INSERT INTO crypto (filename, datetime, type, %s, %s, tx_id, fee) VALUES (?, ?, ?, ?, ?, ?, ?);"
                    % (coin, wallet), (fn[1], date_time, 'xfer', str(qty + fee), str(-qty), tx_id, str(fee)))
        con.commit()
        if "BTC" in coin:
            coin_tx_list[tx_id] = coinObject(Decimal(qty) * SATOSHI, fn[1], coin, coin_addr)
        else:
            coin_tx_list[tx_id] = coinObject(Decimal(qty), fn[1], coin, coin_addr)
        coin_from_exchange.append(tx_id)
    return [coin_tx_list, coin_from_exchange]


def bin_history(fn, cur, con):
    wb = load_workbook(fn[2])
    ws = wb['sheet1']
    # print (ws.min_row)
    # print (ws.max_row)
    previous_time = ""
    previous_date = ""
    previous_market = ""
    coin_qty = 0
    coin_fee_qty = 0
    first_iteration = 1
    coin = ""
    coin_fee = ""
    for row in ws.iter_rows(min_row=2):
        date_time = str(row[0].value).split(" ")
        # print (first_iteration)
        # print (previous_date, date_time[0])
        # print (previous_market, row[1].value)
        if first_iteration == 0 and (previous_date != date_time[0] or previous_market != row[1].value):
            add_column_if_not_in_table(coin, cur, con)
            add_column_if_not_in_table(coin_fee, cur, con)
            cur.execute("INSERT INTO crypto (filename, datetime, type, %s, %s) VALUES (?, ?, ?, ?, ?);" %
                        (coin, coin_fee), (str(fn[1]), str(previous_date) + " " + str(previous_time), 'trade',
                        str(coin_qty), str(coin_fee_qty)))
            con.commit()
            coin_qty = 0
            coin_fee_qty = 0
        first_iteration = 0
        coin = row[1].value
        coin_fee = row[7].value
        tran_type = row[2].value
        amount = row[4].value
        total = row[5].value
        fee = row[6].value
        # print ("coin.find(coin_fee)", str(coin.find(coin_fee)))
        # print ("coin", coin)
        # print ("coin_fee", coin_fee)
        if coin.find(coin_fee) == 0:
            if tran_type == 'BUY':
                coin_qty = coin_qty - Decimal(total)
                coin_fee_qty = coin_fee_qty + Decimal(amount) - Decimal(fee)
            else:
                coin_qty = coin_qty + Decimal(total)
                coin_fee_qty = coin_fee_qty - Decimal(total) - Decimal(fee)
        else:
            if tran_type == 'BUY':
                coin_qty = coin_qty + Decimal(amount)
                coin_fee_qty = coin_fee_qty - Decimal(total) - Decimal(fee)
            else:
                coin_qty = coin_qty - Decimal(amount)
                coin_fee_qty = coin_fee_qty + Decimal(total) - Decimal(fee)
        coin =  str(row[1].value).replace(coin_fee, "") + "_Bin"
        coin_fee =  str(coin_fee) + "_Bin"
        previous_date = date_time[0]
        previous_time = date_time[1]
        previous_market = row[1].value
    add_column_if_not_in_table(coin, cur, con)
    add_column_if_not_in_table(coin_fee, cur, con)
    cur.execute("INSERT INTO crypto (filename, datetime, type, %s, %s) VALUES (?, ?, ?, ?, ?);" % (coin, coin_fee),
                [str(fn[1]), str(previous_date + " " + previous_time), 'trade', str(coin_qty), str(coin_fee_qty)])
    con.commit()

